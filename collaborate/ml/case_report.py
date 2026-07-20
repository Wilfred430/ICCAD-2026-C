"""Per-case constraint/violation report, exported to Excel.

    python -m ml.case_report --weights ml/weights/tree_v2.pt --samples 4

Runs the SAME pipeline as `ml/eval_full.py` (topology sampling -> aspect
sweep -> push_past portfolio -> HPWL nudge) over all 100 validation cases,
but instead of only printing a Total Score, records EVERY hard/soft
constraint outcome and per-case runtime/HPWL/area for every case, then
writes a two-sheet .xlsx:

  * "Per-Case" -- one row per case: feasibility, each hard-violation flag
    (overlap/area/fixed/preplaced), each soft-violation count
    (V_grouping/V_mib/V_boundary/N_soft/V_relative), HPWL_int/ext/gap,
    area_gap, cost, and wall-clock runtime.
  * "Summary"  -- totals/rates across all 100 cases: feasible count, how
    many cases hit each hard violation, summed soft-violation counts,
    total/mean runtime, mean HPWL_gap/area_gap, mean Cost, and the spec's
    e^(n/12)-weighted Total Score.

Re-running overwrites the same file (`--out`, default `case_report.xlsx`)
so it's always a fresh snapshot of the current pipeline's behaviour.
"""

from __future__ import annotations

import argparse
import math
import time

import torch
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from .data import case_to_features
from .model_tree import TreeGenerator
from .pack_tree import build_lc_rc, pack_btree
from .contest_cost import evaluate as evaluate_cost
from .run_pipeline import load_case_raw, build_dims_and_hints
from .eval_full import ASPECT_RATIOS, dims_with_aspect, best_over_aspects, hpwl_nudge


HEADER_FILL = PatternFill(start_color="FFD9E1F2", end_color="FFD9E1F2", fill_type="solid")
BAD_FILL = PatternFill(start_color="FFF8CBAD", end_color="FFF8CBAD", fill_type="solid")
BOLD = Font(bold=True)

PER_CASE_COLS = [
    "case", "n", "runtime_s", "feasible",
    "overlap_violation", "area_violation", "fixed_violation", "preplaced_violation",
    "v_grouping", "v_mib", "v_boundary", "n_soft", "v_relative",
    "hpwl_int", "hpwl_ext", "hpwl_gap_pct", "area_gap_pct", "cost",
]


def run_pipeline_for_report(weights, val, samples, seed, device, limit):
    torch.manual_seed(seed)
    gen = torch.Generator(device=device).manual_seed(seed)

    ckpt = torch.load(weights, map_location=device, weights_only=False)
    cfg = ckpt["config"]
    model = TreeGenerator(hidden_dim=cfg["hidden_dim"], n_ctx_layers=cfg["n_ctx_layers"],
                           n_dec_layers=cfg["n_dec_layers"], n_heads=cfg["n_heads"],
                           max_blocks=cfg["max_blocks"]).to(device)
    model.load_state_dict(ckpt["model_state"])
    model.eval()
    max_n, max_t = cfg["max_blocks"], cfg["max_terms"]

    records = []
    for case_idx in range(min(limit, 100)):
        t0 = time.time()
        blocks, b2b, p2b, pins_pos, metrics, geometry, cfg_name = load_case_raw(val, case_idx)
        n = blocks.shape[0]
        baseline_area = float(metrics[0])
        baseline_hpwl = float(metrics[6]) + float(metrics[7])

        feat = case_to_features(blocks, b2b, p2b, geometry)
        blocks_feat = torch.zeros((1, max_n, feat.shape[1])); blocks_feat[0, :n] = feat
        blocks_mask = torch.zeros((1, max_n), dtype=torch.bool); blocks_mask[0, :n] = True
        t_use = min(pins_pos.shape[0], max_t)
        terms = torch.zeros((1, max_t, 2)); terms[0, :t_use] = pins_pos[:t_use]
        terms_mask = torch.zeros((1, max_t), dtype=torch.bool); terms_mask[0, :t_use] = True

        base_dims, is_preplaced, preplaced_xy = build_dims_and_hints(blocks, geometry)
        cluster_id = {i: int(blocks[i, 4]) for i in range(n)}
        boundary_code = {i: int(blocks[i, 5]) for i in range(n)}

        best_after = None
        best_after_pack = None
        best_after_dims = None

        for _ in range(samples):
            out = model.generate(blocks_feat.to(device), blocks_mask.to(device),
                                  terms.to(device), terms_mask.to(device),
                                  n_blocks=n, temperature=1.0, sample=True, generator=gen)
            root = int(out["gen_order"][0])
            lc, rc = build_lc_rc(root, out["parent_id"], out["direction"], n,
                                 gen_order=out["gen_order"].tolist())

            opt_pack, opt_cc = best_over_aspects(root, lc, rc, blocks, base_dims, is_preplaced,
                                          preplaced_xy, b2b, p2b, pins_pos,
                                          baseline_area, baseline_hpwl, ASPECT_RATIOS,
                                          cluster_id=cluster_id, boundary_code=boundary_code)
            if best_after is None or opt_cc.cost < best_after.cost:
                best_after = opt_cc
                best_after_pack = opt_pack
                best_after_dims = {i: (opt_pack.w[i], opt_pack.h[i]) for i in range(n)}

        if best_after_pack is not None:
            hpwl_nudge(best_after_pack, best_after_dims, blocks, b2b, p2b, pins_pos)
            nudged = evaluate_cost(best_after_pack.x, best_after_pack.y, best_after_dims,
                                   blocks, b2b, p2b, pins_pos, baseline_area, baseline_hpwl)
            if nudged.cost < best_after.cost:
                best_after = nudged

        runtime_s = time.time() - t0
        cc = best_after
        records.append({
            "case": cfg_name, "n": n, "runtime_s": runtime_s, "feasible": cc.feasible,
            "overlap_violation": cc.overlap_violation, "area_violation": cc.area_violation,
            "fixed_violation": cc.fixed_violation, "preplaced_violation": cc.preplaced_violation,
            "v_grouping": cc.v_grouping, "v_mib": cc.v_mib, "v_boundary": cc.v_boundary,
            "n_soft": cc.n_soft, "v_relative": cc.v_relative,
            "hpwl_int": cc.hpwl_int, "hpwl_ext": cc.hpwl_ext, "hpwl_gap_pct": cc.hpwl_gap * 100.0,
            "area_gap_pct": cc.area_gap * 100.0, "cost": cc.cost,
        })
        print(f"  {cfg_name}: n={n:3d} feasible={cc.feasible} cost={cc.cost:.3f} "
              f"Vgrp={cc.v_grouping} Vmib={cc.v_mib} Vbnd={cc.v_boundary} "
              f"runtime={runtime_s:.1f}s")

    return records


def write_report(records, out_path, tau: float = 12.0):
    wb = Workbook()
    ws = wb.active
    ws.title = "Per-Case"

    for col_idx, name in enumerate(PER_CASE_COLS, start=1):
        c = ws.cell(row=1, column=col_idx, value=name)
        c.font = BOLD
        c.fill = HEADER_FILL

    hard_cols = {"overlap_violation", "area_violation", "fixed_violation", "preplaced_violation"}
    for row_idx, rec in enumerate(records, start=2):
        for col_idx, name in enumerate(PER_CASE_COLS, start=1):
            val = rec[name]
            if isinstance(val, bool):
                val = "Yes" if val else "No"
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            if name == "feasible" and rec[name] is False:
                cell.fill = BAD_FILL
            if name in hard_cols and rec[name] is True:
                cell.fill = BAD_FILL
            if name in ("v_grouping", "v_mib", "v_boundary") and rec[name] > 0:
                cell.fill = BAD_FILL
        ws.cell(row=row_idx, column=1).font = Font()

    # --- Totals row, directly under the case rows (row n_cases+2, e.g. row
    # 102 for 100 cases) so avg runtime / avg cost / total soft violations
    # are visible without switching to the Summary sheet. ---
    n_cases = len(records)
    total_vgrp = sum(r["v_grouping"] for r in records)
    total_vmib = sum(r["v_mib"] for r in records)
    total_vbnd = sum(r["v_boundary"] for r in records)
    total_nsoft = sum(r["n_soft"] for r in records)
    total_runtime = sum(r["runtime_s"] for r in records)
    mean_runtime = total_runtime / n_cases if n_cases else 0.0
    mean_cost = sum(r["cost"] for r in records) / n_cases if n_cases else 0.0

    totals_row = n_cases + 2
    totals = {
        "case": "AVG / TOTAL",
        "runtime_s": round(mean_runtime, 3),
        "n_soft": total_nsoft,
        "v_grouping": total_vgrp,
        "v_mib": total_vmib,
        "v_boundary": total_vbnd,
        "cost": round(mean_cost, 4),
    }
    for col_idx, name in enumerate(PER_CASE_COLS, start=1):
        if name in totals:
            cell = ws.cell(row=totals_row, column=col_idx, value=totals[name])
            cell.font = BOLD
            cell.fill = HEADER_FILL

    for col_idx, name in enumerate(PER_CASE_COLS, start=1):
        width = max(10, len(name) + 2)
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.freeze_panes = "A2"

    # --- Summary sheet ---
    ws2 = wb.create_sheet("Summary")
    n_feasible = sum(1 for r in records if r["feasible"])
    n_overlap = sum(1 for r in records if r["overlap_violation"])
    n_area_v = sum(1 for r in records if r["area_violation"])
    n_fixed_v = sum(1 for r in records if r["fixed_violation"])
    n_preplaced_v = sum(1 for r in records if r["preplaced_violation"])
    mean_hpwl_gap = sum(r["hpwl_gap_pct"] for r in records) / n_cases if n_cases else 0.0
    mean_area_gap = sum(r["area_gap_pct"] for r in records) / n_cases if n_cases else 0.0
    weights = [math.exp(r["n"] / tau) for r in records]
    total_score = (sum(c * w for c, w in zip((r["cost"] for r in records), weights))
                   / sum(weights)) if weights else 0.0

    summary_rows = [
        ("Total cases", n_cases),
        ("Feasible cases", f"{n_feasible}/{n_cases}"),
        ("Infeasible cases", n_cases - n_feasible),
        ("-- Hard constraint violations (case count) --", ""),
        ("Overlap violation", n_overlap),
        ("Area violation (soft-block 1% tolerance)", n_area_v),
        ("Fixed-shape violation", n_fixed_v),
        ("Preplaced violation", n_preplaced_v),
        ("-- Soft constraint violations (summed across all cases) --", ""),
        ("Total V_grouping", total_vgrp),
        ("Total V_mib", total_vmib),
        ("Total V_boundary", total_vbnd),
        ("Total N_soft (sum of all three)", total_nsoft),
        ("-- Runtime --", ""),
        ("Total runtime (s)", round(total_runtime, 2)),
        ("Mean runtime per case (s)", round(mean_runtime, 3)),
        ("-- Quality --", ""),
        ("Mean HPWL_gap (%)", round(mean_hpwl_gap, 2)),
        ("Mean area_gap (%)", round(mean_area_gap, 2)),
        ("Mean Cost (unweighted)", round(mean_cost, 4)),
        (f"Total Score (e^(n/{tau:.0f}) weighted)", round(total_score, 4)),
    ]
    for row_idx, (label, val) in enumerate(summary_rows, start=1):
        lc = ws2.cell(row=row_idx, column=1, value=label)
        if label.startswith("--"):
            lc.font = BOLD
        ws2.cell(row=row_idx, column=2, value=val)
    ws2.column_dimensions["A"].width = 45
    ws2.column_dimensions["B"].width = 16

    wb.save(out_path)
    print(f"\n[case_report] wrote {out_path}")
    print(f"[case_report] feasible {n_feasible}/{n_cases}  Total Score={total_score:.4f}  "
          f"Vgrp={total_vgrp} Vmib={total_vmib} Vbnd={total_vbnd}")


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--weights", default="ml/weights/tree_v2.pt")
    ap.add_argument("--val", default="d:/ICCAD-2026-C/ICCAD-C-FloorSet-official/LiteTensorDataTest")
    ap.add_argument("--samples", type=int, default=4)
    ap.add_argument("--seed", type=int, default=0)
    ap.add_argument("--device", default="cuda" if torch.cuda.is_available() else "cpu")
    ap.add_argument("--limit", type=int, default=100)
    ap.add_argument("--out", default="case_report.xlsx")
    args = ap.parse_args()

    print(f"[case_report] loaded {args.weights}  device={args.device}  samples={args.samples}")
    records = run_pipeline_for_report(args.weights, args.val, args.samples, args.seed,
                                       args.device, args.limit)
    write_report(records, args.out)


if __name__ == "__main__":
    main()
